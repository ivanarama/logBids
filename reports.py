from datetime import date
import asyncio, ssl, smtplib
from email.message import EmailMessage
from email.utils import formataddr
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy import cast, Date
from models import SessionLocal, Bid
from config import settings
from collections import defaultdict
from collections import OrderedDict

async def generate_and_send_report(report_date: date | None = None, debug: bool = True):
    db = SessionLocal()
    try:
        if report_date is None:
            report_date = date.today()

        rows = db.query(
            Bid.branch, Bid.direction, Bid.bidid, Bid.biddate, Bid.created_at, Bid.isrepeat
        ).filter(cast(Bid.biddate, Date) == report_date).order_by(
            Bid.branch, Bid.direction, Bid.biddate
        ).all()

    finally:
        db.close()

    def _dbg(msg: str):
        if debug:
            print(msg)

    _dbg(f"[REPORT] report_date={report_date} rows_loaded={len(rows)}")

    if not rows:
        _dbg("[REPORT] Нет заявок для выбранной даты.")
        return None

    # --- Собираем дерево branch -> direction -> bids ---
    tree = OrderedDict()
    branch_stats = defaultdict(lambda: {"total": 0, "repeat": 0})
    direction_stats = defaultdict(lambda: {"total": 0, "repeat": 0})

    for branch, direction, bidid, biddate, created_at, isRepeat in rows:
        tree.setdefault(branch, OrderedDict())
        tree[branch].setdefault(direction, []).append((bidid, biddate, created_at, isRepeat))
        branch_stats[branch]["total"] += 1
        branch_stats[branch]["repeat"] += 1 if isRepeat else 0
        direction_stats[(branch, direction)]["total"] += 1
        direction_stats[(branch, direction)]["repeat"] += 1 if isRepeat else 0

    _dbg(f"[STATS] branches={len(tree)} directions={sum(len(d) for d in tree.values())}")
    for b, v in branch_stats.items():
        _dbg(f"  BRANCH {b}: total={v['total']} repeat={v['repeat']}")
    for (b, d), v in direction_stats.items():
        _dbg(f"  DIR {b} / {d}: total={v['total']} repeat={v['repeat']}")

    # --- Пишем Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    ws.append(["Филиал", "Направление", "BidID", "BidDate", "CreatedAt", "Платные", "Повторные", "Всего"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    group_info = []

    # Кастомный порядок филиалов
    branch_order = {
        "МСК": 1,
        "СПБ": 2,
        "ННОВ": 3,
        "РнД": 4,
        "КРД": 5,
        "ВРН": 6,
        "ЕКБ": 7,
        "НСК": 8,
        "ЛПЦ": 9,
        "КЗН": 10,
        "САМ": 11,
        "УФА": 12,
        "ОМС": 13,
        "КРЯ": 14,
        "ПРМ": 15,
        "ВГГ": 16,
        "ТМН": 17,
        "СРТ": 18,
        "ЧЛБ": 19,
    }
    sorted_branches = sorted(tree.keys(), key=lambda b: (branch_order.get(b, 10_000), str(b)))

    for branch in sorted_branches:
        directions = tree[branch]
        # branch header
        b_total = branch_stats[branch]["total"]
        b_repeat = branch_stats[branch]["repeat"]
        b_paid = b_total - b_repeat

        branch_header_row = row_idx
        ws.append([branch, "", "", "", "", b_paid, b_repeat, b_total])
        _dbg(f"[NEW BRANCH] {branch} header_row={branch_header_row}")
        row_idx += 1

        for direction, bids in directions.items():
            d_total = direction_stats[(branch, direction)]["total"]
            d_repeat = direction_stats[(branch, direction)]["repeat"]
            d_paid = d_total - d_repeat

            direction_header_row = row_idx
            ws.append(["", direction, "", "", "", d_paid, d_repeat, d_total])
            _dbg(f"  [NEW DIRECTION] {direction} header_row={direction_header_row}")
            row_idx += 1

            # заявки
            first_child = row_idx
            for bidid, biddate, created_at, isRepeat in bids:
                biddate_s = biddate.strftime("%d.%m.%Y %H:%M:%S") if biddate else ""
                created_s = created_at.strftime("%d.%m.%Y %H:%M:%S") if created_at else ""
                paid_val = 0 if isRepeat else 1
                repeat_val = 1 if isRepeat else 0
                ws.append(["", "", bidid, biddate_s, created_s, paid_val, repeat_val, 1])
                row_idx += 1
            last_child = row_idx - 1

            if last_child >= first_child:

                # defer grouping; apply after building all rows to preserve highest outline level
                _dbg(f"    [GROUP bids] {branch} / {direction}: rows {first_child}-{last_child} (level=3, hidden=True)")
                group_info.append(("direction-bids", branch, direction, first_child, last_child))
                #print(f"    [GROUP bids] {branch} / {direction}: rows {first_child}-{last_child} (hidden=True)")

                
                _dbg(f"    [GROUP direction] {branch} / {direction}: rows {direction_header_row + 1}-{last_child} (level=2, hidden=False)")
                group_info.append(("direction", branch, direction, direction_header_row + 1, last_child))
                #print(f"    [GROUP direction] {branch} / {direction}: rows {direction_header_row + 1}-{last_child} (hidden=False)")

        # филиал (уровень 1, раскрыт)
        br_first_child = branch_header_row + 1
        br_last_child = row_idx - 1
        if br_last_child >= br_first_child:
            group_info.append(("branch", branch, None, br_first_child, br_last_child))
            _dbg(f"[GROUP branch] {branch}: rows {br_first_child}-{br_last_child} (hidden=False)")

    ws.sheet_properties.outlinePr.summaryBelow = False
    ws.sheet_properties.outlinePr.summaryRight = True
    ws.sheet_properties.outlinePr.applyStyles = True
    ws.sheet_view.showOutlineSymbols = True

    # --- дамп структуры ---
    _dbg("[REPORT] group_info:")
    for g in group_info:
        if g[0] == "branch":
            _dbg(f"  branch {g[1]}: rows {g[3]}-{g[4]}")
        elif g[0] == "direction":
            _dbg(f"  direction {g[1]} / {g[2]} (with bids): rows {g[3]}-{g[4]}")
        else:
            _dbg(f"  bids {g[1]} / {g[2]}: rows {g[3]}-{g[4]}")

    # Apply groups in ascending order (1 → 2 → 3) so the deepest level wins on overlap
    for level in (1, 2, 3):
        for g in group_info:
            gtype, gbranch, gdir, gstart, gend = g
            if level == 1 and gtype == "branch":
                ws.row_dimensions.group(gstart, gend, outline_level=1, hidden=False)
            elif level == 2 and gtype == "direction":
                ws.row_dimensions.group(gstart, gend, outline_level=2, hidden=False)
            elif level == 3 and gtype == "direction-bids":
                ws.row_dimensions.group(gstart, gend, outline_level=3, hidden=True)

    last_row = row_idx - 1
    _dbg("=== OUTLINE LEVELS (row -> outlineLevel | col1 | col2) ===")
    for r in range(1, last_row + 1):
        ol = getattr(ws.row_dimensions[r], "outlineLevel", 0)
        c1 = ws.cell(row=r, column=1).value
        c2 = ws.cell(row=r, column=2).value
        _dbg(f"row {r}: outlineLevel={ol} | col1={c1!r} | col2={c2!r}")

    # автоширина
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
        ws.column_dimensions[col_letter].width = max_len + 2

    filename = f"/tmp/report_{report_date}.xlsx"
    wb.save(filename)
    _dbg(f"[REPORT] Отчёт сохранён: {filename}")

    # --- отправка почтой только если debug=False ---
    if debug:
        _dbg("[REPORT] debug=True — отправка почты пропущена")
        return filename

    email_addresses = [e.strip() for e in settings.REPORT_EMAIL_TO.split(",") if e.strip()]
    if not email_addresses:
        print("[REPORT] Нет email-адресов для отчёта")
        return filename

    def _send():
        msg = EmailMessage()
        msg["Subject"] = f"Ежедневный отчёт {report_date}"
        msg["From"] = formataddr(("Система отчётов", settings.SMTP_USER))
        msg["To"] = ", ".join(email_addresses)
        msg.set_content("В приложении отчёт", subtype="plain", charset="utf-8")
        with open(filename, "rb") as f:
            msg.add_attachment(f.read(),
                               maintype="application",
                               subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               filename=("report_" + str(report_date) + ".xlsx"))
        ctx = ssl.create_default_context()
        with smtplib.SMTP_SSL(settings.SMTP_SERVER, int(settings.SMTP_PORT), context=ctx, timeout=30) as server:
            server.login(settings.SMTP_USER, settings.SMTP_PASS)
            server.send_message(msg)
            _dbg("[MAIL] письмо отправлено")

    await asyncio.to_thread(_send)
    return filename