from fastapi import FastAPI, Header, HTTPException
from pydantic import BaseModel, field_validator
from datetime import datetime
from sqlalchemy import create_engine, Column, Integer, String, DateTime, func
from sqlalchemy.orm import declarative_base, sessionmaker
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from apscheduler.schedulers.background import BackgroundScheduler
from config import settings
import asyncio
import ssl
from email.message import EmailMessage
from email.utils import formataddr
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from sqlalchemy import Date
# --- DB setup ---
engine = create_engine(settings.DATABASE_URL)
SessionLocal = sessionmaker(bind=engine)
Base = declarative_base()

app = FastAPI()

class Bid(Base):
    __tablename__ = "bids"
    id = Column(Integer, primary_key=True, index=True)
    bidid = Column(String, index=True)
    biddate = Column(DateTime)
    direction = Column(String)
    branch = Column(String)
    source_id = Column(String, nullable=False)
    created_at = Column(DateTime(timezone=True), server_default=func.now())

Base.metadata.create_all(bind=engine)

# --- Pydantic model ---
class BidRequest(BaseModel):
    bidid: str
    biddate: datetime
    direction: str
    branch: str
    source_id: str 
    @field_validator("biddate", mode="before")
    def parse_biddate(cls, v):
        if isinstance(v, str):
            try:
                return datetime.strptime(v, "%d.%m.%Y %H:%M:%S")
            except ValueError:
                raise ValueError("biddate must be in format DD.MM.YYYY HH:MM:SS")
        return v

# --- API add_bid ---
@app.post("/add_bid/")
async def add_bid(bid: BidRequest, authorization: str = Header(...)):
    if authorization != settings.SECRET_KEY:
        raise HTTPException(status_code=401, detail="Unauthorized")
    db = SessionLocal()
    new_bid = Bid(
        bidid=bid.bidid,
        biddate=bid.biddate,
        direction=bid.direction,
        branch=bid.branch,
        source_id=bid.source_id   # сохраняем в БД
        # created_at ставится автоматически
    )
    db.add(new_bid)
    db.commit()
    db.refresh(new_bid)
    db.close()
    return {"status": "ok", "id": new_bid.id}

# --- Отчёт ---
async def generate_and_send_report():
    today = date.today()
    db = SessionLocal()
    rows = db.query(
        Bid.branch, Bid.direction, Bid.bidid, Bid.biddate, Bid.created_at
    ).filter(Bid.created_at.cast(Date) == today).order_by(
        Bid.branch, Bid.direction, Bid.biddate
    ).all()
    db.close()

    if not rows:
        print("Нет заявок за сегодня")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Отчёт"

    ws.append(["Филиал", "Направление", "Номер", "Дата", "Запись сделана"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row_idx = 2
    current_branch = None
    current_direction = None
    branch_count = 0
    direction_count = 0
    branch_start_idx = row_idx
    direction_start_idx = row_idx

    # Считаем кол-во заранее
    from collections import defaultdict
    branch_counts = defaultdict(int)
    direction_counts = defaultdict(int)
    for branch, direction, *_ in rows:
        branch_counts[branch] += 1
        direction_counts[(branch, direction)] += 1

    for branch, direction, bidid, biddate, created_at in rows:
        # Branch
        if branch != current_branch:
            current_branch = branch
            branch_count = branch_counts[branch]
            ws.append([f"{branch} ({branch_count})"])
            ws[row_idx][0].font = Font(bold=True)
            row_idx += 1

        # Direction
        if direction != current_direction:
            current_direction = direction
            direction_count = direction_counts[(branch, direction)]
            ws.append(["", f"{direction} ({direction_count})"])
            ws[row_idx][1].font = Font(italic=True)
            row_idx += 1

        # Заявка
        ws.append(["", "", bidid, biddate.strftime("%d.%m.%Y %H:%M:%S"), created_at.strftime("%d.%m.%Y %H:%M:%S")])
        row_idx += 1

    # Автоширина колонок
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Сохраняем файл
    filename = "/tmp/report.xlsx"
    wb.save(filename)

    # список получателей из .env (через запятую)
    email_addresses = [e.strip() for e in settings.REPORT_EMAIL_TO.split(",") if e.strip()]
    if not email_addresses:
        print("Нет email-адресов для отправки")
        return

    def _send():
        msg = EmailMessage()
        msg["Subject"] = "Ежедневный отчёт"
        msg["From"] = formataddr(("А-Айсберг", settings.SMTP_USER))
        msg["To"] = ", ".join(email_addresses)
        msg.set_content("В приложении ежедневный отчёт", subtype="plain", charset="utf-8")

        # прикрепляем Excel
        with open(filename, "rb") as f:
            file_data = f.read()
            msg.add_attachment(
                file_data,
                maintype="application",
                subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                filename="report.xlsx"
            )

        try:
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(settings.SMTP_SERVER, int(settings.SMTP_PORT), timeout=30, context=ctx) as server:
                server.ehlo()
                server.login(settings.SMTP_USER, settings.SMTP_PASS)  # пароль приложения Gmail
                server.send_message(msg)
                print(f"Письмо отправлено на: {', '.join(email_addresses)}")
        except smtplib.SMTPAuthenticationError:
            print("Ошибка аутентификации: проверь пароль приложения Gmail")

    # неблокирующий вызов
    await asyncio.to_thread(_send)

# --- ручной запуск отчёта ---
@app.post("/send_report_now/")
async def  send_report_now(authorization: str = Header(...)):
    if authorization != settings.SECRET_KEY:
        raise HTTPException(status_code=401, detail="Unauthorized")
    await generate_and_send_report()
    return {"status": "ok", "message": "report sent"}

# --- планировщик ---
scheduler = BackgroundScheduler()
scheduler.add_job(generate_and_send_report, "cron", hour=23, minute=59)
scheduler.start()